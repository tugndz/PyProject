import os
import time
import xml.etree.ElementTree as ET
import pandas as pd
from urllib.parse import urlparse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

# Hàm thiết lập trình duyệt Chrome
def setup_driver(download_dir):
    """
    Thiết lập trình duyệt Chrome với các tùy chọn tải xuống và khởi tạo WebDriverWait
    Args:
        download_dir: Thư mục lưu file XML tải về
    Returns:
        Tuple chứa đối tượng webdriver và WebDriverWait
    """
    os.makedirs(download_dir, exist_ok=True)  # Tạo thư mục tải xuống nếu chưa tồn tại
    options = Options()
    options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,  # Thiết lập thư mục tải xuống
        "download.prompt_for_download": False,  # Tắt hộp thoại hỏi nơi lưu file
        "plugins.always_open_pdf_externally": True,  # Tự động tải PDF (nếu có)
        "safebrowsing.enabled": True  # Bật chế độ bảo mật
    })
    driver = webdriver.Chrome(options=options)  # Khởi tạo trình duyệt
    return driver, WebDriverWait(driver, 10)  # Trả về driver và wait với timeout 10 giây

# Hàm tra cứu và tải file XML
def tra_cuu_va_tai_xml(driver, wait, mst, mtc, url, download_dir):
    """
    Tra cứu hóa đơn trên các trang web hỗ trợ và tải file XML
    Args:
        driver: Đối tượng webdriver
        wait: Đối tượng WebDriverWait
        mst: Mã số thuế bên bán
        mtc: Mã tra cứu hóa đơn
        url: URL trang tra cứu
        download_dir: Thư mục lưu file XML
    Returns:
        Đường dẫn file XML tải về hoặc None nếu thất bại
    """
    driver.get(url)  # Mở trang tra cứu
    domain = urlparse(url).netloc  # Lấy domain từ URL

    try:
        # Xử lý cho trang fpt.com.vn
        if "fpt.com.vn" in domain:
            wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='MST bên bán']"))).send_keys(mst)  # Nhập MST
            driver.find_element(By.XPATH, "//input[@placeholder='Mã tra cứu hóa đơn']").send_keys(mtc)  # Nhập mã tra cứu
            driver.find_element(By.XPATH, "//button[contains(@class, 'webix_button') and contains(text(), 'Tra cứu')]").click()  # Nhấn nút tra cứu
            time.sleep(2)  # Đợi kết quả
            driver.implicitly_wait(10)  # Đợi ngầm 10 giây
            driver.find_element(By.XPATH, "//button[span[contains(@class, 'mdi-xml')] and contains(text(), 'Tải XML')]").click()  # Tải XML
            time.sleep(0.1)  # Đợi ngắn để bắt đầu tải

        # Xử lý cho trang meinvoice.vn
        elif "meinvoice.vn" in domain:
            wait.until(EC.presence_of_element_located((By.NAME, "txtCode"))).send_keys(mtc)  # Nhập mã tra cứu
            driver.find_element(By.ID, "btnSearchInvoice").click()  # Nhấn nút tra cứu
            time.sleep(2)  # Đợi kết quả
            driver.find_element(By.CLASS_NAME, "download").click()  # Mở menu tải
            time.sleep(1)  # Đợi menu xuất hiện
            driver.find_element(By.CLASS_NAME, "txt-download-xml").click()  # Tải XML

        # Xử lý cho trang ehoadon.vn
        elif "ehoadon.vn" in domain:
            wait.until(EC.presence_of_element_located((By.ID, "txtInvoiceCode"))).send_keys(mtc)  # Nhập mã tra cứu
            driver.find_element(By.CLASS_NAME, "btnSearch").click()  # Nhấn nút tra cứu
            wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "frameViewInvoice")))  # Chuyển sang iframe
            time.sleep(2)  # Đợi kết quả
            driver.find_element(By.ID, "btnDownload").click()  # Mở menu tải
            time.sleep(1)  # Đợi menu xuất hiện
            driver.execute_script("document.querySelector('#divDownloads .dropdown-menu').style.display='block';")  # Hiển thị menu tải
            driver.find_element(By.ID, "LinkDownXML").click()  # Tải XML

        else:
            print(f"Không hỗ trợ URL: {url}")
            return None 

        # Chờ file XML xuất hiện trong thư mục tải xuống
        start_time = time.time()
        timeout = 20
        while time.time() - start_time < timeout:
            files = [f for f in os.listdir(download_dir) if f.endswith(".xml")]  # Lấy danh sách file XML
            if files:
                files = sorted(files, key=lambda x: os.path.getmtime(os.path.join(download_dir, x)), reverse=True)  # Sắp xếp theo thời gian mới nhất
                return os.path.join(download_dir, files[0])  # Trả về file mới nhất
            time.sleep(1)  # Đợi 1 giây trước khi kiểm tra lại

    except Exception as e:
        print(f"Lỗi khi xử lý {url}: {e}")  # Ghi lỗi nếu có
        return None
    return None

# Hàm đọc thông tin từ file XML
def read_xml_info(xml_path):
    """
    Đọc thông tin từ file XML và trả về dictionary chứa dữ liệu hóa đơn
    Args:
        xml_path: Đường dẫn tới file XML
    Returns:
        Dictionary chứa thông tin hóa đơn hoặc None nếu thất bại
    """
    try:
        # Phân tích file XML
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # Tìm node chính chứa thông tin hóa đơn
        hoa_don_node = root.find(".//HDon")
        invoice_node = hoa_don_node.find("DLHDon") if hoa_don_node is not None else None
        if invoice_node is None:
            for tag in [".//DLHDon", ".//TDiep", ".//Invoice"]:  # Thử các tag khác
                node = root.find(tag)
                if node is not None:
                    invoice_node = node
                    break
            else:
                return None  # Không tìm thấy node chính

        # Hàm hỗ trợ tìm giá trị theo đường dẫn XML
        def find(path):
            current = invoice_node
            for part in path.split("/"):
                if current is not None:
                    current = current.find(part)
                else:
                    return None
            return current.text if current is not None else None

        # Tìm số tài khoản bán hàng
        stk_ban_hang = find("NDHDon/NBan/STKNHang")
        if not stk_ban_hang:
            for thongtin in invoice_node.findall(".//NBan/TTKhac/TTin"):
                if thongtin.findtext("TTruong") == "SellerBankAccount":
                    stk_ban_hang = thongtin.findtext("DLieu")
                    break

        # Trả về dictionary chứa thông tin hóa đơn
        return {
            'Số hóa đơn': find("TTChung/SHDon"),
            'Đơn vị bán hàng': find("NDHDon/NBan/Ten"),
            'Mã số thuế bán': find("NDHDon/NBan/MST"),
            'Địa chỉ bán': find("NDHDon/NBan/DChi"),
            'Số tài khoản bán': stk_ban_hang,
            'Họ tên người mua hàng': find("NDHDon/NMua/Ten"),
            'Địa chỉ mua': find("NDHDon/NMua/DChi"),
            'Mã số thuế mua': find("NDHDon/NMua/MST"),
        }
    except Exception as e:
        print(f"Lỗi đọc XML: {e}")
        return None

# Hàm ghi dữ liệu ra file Excel
def write_excel(filepath, data):
    """
    Ghi dữ liệu vào file Excel, tạo mới nếu chưa tồn tại
    Args:
        filepath: Đường dẫn file Excel
        data: Danh sách dữ liệu để ghi
    """
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "STT", "MST", "Mã tra cứu", "URL",
            "Số hóa đơn", "Đơn vị bán hàng",
            "Mã số thuế bán", "Địa chỉ bán",
            "Số tài khoản bán", "Họ tên người mua hàng",
            "Địa chỉ mua", "Mã số thuế mua", "Ghi chú"
        ])  # Tạo header cho file mới
    else:
        wb = load_workbook(filepath)  # Mở file Excel hiện có
        ws = wb.active
    for row in data:
        ws.append(row)  # Thêm từng dòng dữ liệu
    wb.save(filepath)  # Lưu file

# Hàm chính điều khiển luồng chương trình
def main():
    """
    Hàm chính: Đọc dữ liệu từ file Excel đầu vào, tra cứu, tải XML và ghi kết quả ra file Excel
    """
    input_file = "input.xlsx"  # File đầu vào chứa MST, mã tra cứu, URL
    output_file = "output.xlsx"  # File đầu ra lưu kết quả
    script_dir = os.path.dirname(os.path.abspath(__file__))  # Thư mục chứa script
    download_dir = os.path.join(script_dir, "down_invoices")  # Thư mục lưu XML

    driver, wait = setup_driver(download_dir)  # Thiết lập trình duyệt
    df = pd.read_excel(input_file, dtype=str)  # Đọc file đầu vào
    result_rows = []  # Danh sách lưu kết quả

    # Duyệt qua từng dòng trong file Excel
    for i, row in df.iterrows():
        mst = str(row.get("Mã số thuế", "")).strip()  # Lấy MST
        mtc = str(row.get("Mã tra cứu", "")).strip()  # Lấy mã tra cứu
        url = str(row.get("URL", "")).strip()  # Lấy URL
        print(f"\n>> Đang xử lý: {mtc} | Trang: {url}")
        xml_file = tra_cuu_va_tai_xml(driver, wait, mst, mtc, url, download_dir)  # Tra cứu và tải XML

        if xml_file:
            info = read_xml_info(xml_file)  # Đọc thông tin từ XML
            if info:
                result_rows.append([
                    i + 1, mst, mtc, url,
                    info["Số hóa đơn"], info["Đơn vị bán hàng"], 
                    info["Mã số thuế bán"], info["Địa chỉ bán"], 
                    info["Số tài khoản bán"], info["Họ tên người mua hàng"], 
                    info["Địa chỉ mua"], info["Mã số thuế mua"], ""
                ])  # Thêm dòng dữ liệu thành công
            else:
                result_rows.append([i + 1, mst, mtc, url] + [""] * 9 + ["Không đọc được XML"])  # Ghi lỗi nếu không đọc được XML
        else:
            result_rows.append([i + 1, mst, mtc, url] + [""] * 9 + ["Tải XML thất bại"])  # Ghi lỗi nếu không tải được XML

    write_excel(output_file, result_rows)  # Ghi kết quả ra file Excel
    time.sleep(5)  # Đợi 5 giây trước khi đóng trình duyệt
    driver.quit()  # Đóng trình duyệt
    print(f"\n==> Đã lưu kết quả vào: {output_file}")

if __name__ == "__main__":
    main()